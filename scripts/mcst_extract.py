# -*- coding: utf-8 -*-
"""
문체부 '2025 전국 문화기반시설 총람'(기준일 2025-01-01, 공공누리 제1유형)에서
서울특별시 시설을 추출하고 Nominatim으로 지오코딩한다.
출력: data/mcst_seoul.json
"""
import io, sys, json, re, time, urllib.parse, urllib.request
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
BASE = r"C:\Users\user\Desktop\관장서류추가\seoul-culture-3d"

wb = openpyxl.load_workbook(f"{BASE}/data/mcst_2025_facilities.xlsx", read_only=True)
sheets = {sn.strip(): sn for sn in wb.sheetnames}

def rows_of(sheet_key):
    return list(wb[sheets[sheet_key]].iter_rows(values_only=True))

def s(v):
    return str(v).strip() if v is not None else ""

items = []

def add(name, gu, addr, typ, sub, oper, pub, year, web, phone):
    name, addr = s(name), s(addr)
    if not name or not addr:
        return
    items.append({
        "시설명": name, "자치구명": s(gu), "주소": addr, "유형": typ, "세부장르": sub,
        "운영주체": s(oper) or "자료 없음", "공공민간": pub,
        "설립연도": s(year)[:10] or "자료 없음", "웹사이트": s(web) or "자료 없음",
        "전화": s(phone) or "자료 없음",
    })

def pubmap(v):
    v = s(v)
    if any(k in v for k in ("국립",)): return "공공"
    if any(k in v for k in ("공립", "시립", "구립", "교육청", "지자체", "시", "구")): return "공공"
    if any(k in v for k in ("사립", "법인", "개인")): return "민간"
    if "대학" in v: return "민간(대학)"
    return "확인 필요"

# 공공도서관: 연번0 시도1 시군구2 설립주체3 도서관명4 주소5 연락처6 홈페이지7 개관년도8
for r in rows_of("공공도서관")[4:]:
    if r and "서울" in s(r[1]):
        add(r[4], r[2], r[5], "공공도서관", "총람: 공공도서관", r[3], pubmap(r[3]), r[8], r[7], r[6])

# 박물관/미술관: 연번0 시도1 시군구2 구분3(국공사립) 종별4 명5 주소6 연락처7 개관8
for key, typ in (("박물관", "박물관"), ("미술관", "미술관")):
    for r in rows_of(key)[6:]:
        if r and "서울" in s(r[1]):
            t = "미술관" if ("미술관" in s(r[5]) or typ == "미술관") else "박물관"
            add(r[5], r[2], r[6], t, f"총람: {typ}({s(r[3])}·{s(r[4])})", r[3], pubmap(r[3]), r[8], "", r[7])

# 생활문화센터: 번호0 시도1 시군구2 명3 운영방식4 운영주체5 공간유형6 개관7 주소8 전화9
for r in rows_of("생활문화센터")[4:]:
    if r and "서울" in s(r[1]):
        add(r[3], r[2], r[8], "복합문화공간", "총람: 생활문화센터", r[5], "공공", r[7], "", r[9])

# 문예회관: 연번0 시도1 시군구2 건립주체3 시설명4 주소5 연락처6 운영주체7 홈페이지8 개관9
for r in rows_of("문예회관")[4:]:
    if r and "서울" in s(r[1]):
        add(r[4], r[2], r[5], "공연장", "총람: 문예회관", r[7] or r[3], "공공", r[9], r[8], r[6])

# 지방문화원: 연번0 시도1 시군구2 문화원명3 원장4 설립일5 주소6 연락처7 홈페이지8
for r in rows_of("지방문화원")[4:]:
    if r and "서울" in s(r[1]):
        add(r[3], r[2], r[6], "문화원·지역문화센터", "총람: 지방문화원", r[3], "공공", r[5], r[8], r[7])

# 문화의집: 번호0 시도1 시군구2 명3 주소4 전화5 홈페이지6 개관일7 운영형태8
for r in rows_of("문화의집")[4:]:
    if r and "서울" in s(r[1]):
        add(r[3], r[2], r[4], "문화원·지역문화센터", "총람: 문화의집", r[8], "공공", r[7], r[6], r[5])

# 문학관: 연번0 시·도1 시군구2 국공사립3 설립주체4 운영방식5 등록번호6 등록일자7 등록형태8 명9 주소10 연락처11
for r in rows_of("문학관")[6:]:
    if r and "서울" in s(r[1]):
        add(r[9], r[2], r[10], "기념관·기념물", "총람: 등록문학관", r[4], pubmap(r[3]), r[7], "", r[11])

# 지역문화재단: 번호0 시도1 시군구2 재단명3 주소4 연락처5 홈페이지6 설립일7
for r in rows_of("(부록)지역문화재단")[4:]:
    if r and "서울" in s(r[1]):
        add(r[3], r[2], r[4], "문화원·지역문화센터", "총람: 지역문화재단", r[3], "공공", r[7], r[6], r[5])

print("총람 서울 시설 추출:", len(items))

# ---------- Nominatim 지오코딩 (1 req/sec 정책 준수) ----------
UA = "seoul-culture-3d-research/1.0 (academic research; contact joyof15@gmail.com)"

def geocode(q):
    url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
        {"q": q, "format": "json", "limit": 1, "countrycodes": "kr"})
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            j = json.loads(resp.read().decode())
            if j:
                return float(j[0]["lon"]), float(j[0]["lat"])
    except Exception as e:
        print("  geocode err:", e)
    return None

def clean_addr(a):
    a = re.sub(r"\([^)]*\)", " ", a)           # 괄호 제거
    a = re.sub(r"\d+층.*$|지하\s?\d.*$", " ", a)  # 층 정보 제거
    return re.sub(r"\s+", " ", a).strip()

# 기존 OSM 시설과 이름이 일치하면 지오코딩 생략(병합 단계에서 OSM 좌표 사용)
import json as _json
existing = set()
try:
    fx = _json.load(open(f"{BASE}/data/facilities.geojson", encoding="utf-8"))
    existing = {re.sub(r"\s+", "", f["properties"]["시설명"]) for f in fx["features"]}
except Exception as e:
    print("기존 시설 로드 실패:", e)

ok = skipped = 0
for i, it in enumerate(items):
    if re.sub(r"\s+", "", it["시설명"]) in existing:
        skipped += 1
        print(f"{i+1}/{len(items)} {it['시설명']} SKIP(기존 일치)")
        continue
    addr = clean_addr(it["주소"])
    cand = [addr]
    m = re.search(r"^(.*?(?:로|길)\s?\d+(?:-\d+)?)", addr)   # 도로명+번호까지만
    if m and m.group(1) != addr:
        cand.append(m.group(1))
    cand.append("서울 " + it["시설명"])                        # 최후: 시설명 검색
    coords = None
    for q in cand:
        coords = geocode(q)
        time.sleep(1.1)
        if coords:
            break
    if coords:
        # 대전 영역 검증
        if 126.7 < coords[0] < 127.3 and 37.3 < coords[1] < 37.8:
            it["경도"], it["위도"] = round(coords[0], 6), round(coords[1], 6)
            ok += 1
        else:
            print(f"  [{it['시설명']}] 좌표가 서울 밖 → 제외 표시")
    print(f"{i+1}/{len(items)} {it['시설명']} {'OK' if it.get('경도') else 'FAIL'}")

print(f"지오코딩 성공: {ok} / 기존일치 생략: {skipped} / 전체: {len(items)}")
with open(f"{BASE}/data/mcst_seoul.json", "w", encoding="utf-8") as f:
    json.dump({"meta": {"source": "문화체육관광부 2025 전국 문화기반시설 총람",
                        "기준일": "2025-01-01", "라이선스": "공공누리 제1유형",
                        "지오코딩": "Nominatim(OSM) — 좌표는 근사치일 수 있음"},
               "items": items}, f, ensure_ascii=False, indent=1)
print("저장: data/mcst_seoul.json")
